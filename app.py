from pathlib import Path
from datetime import datetime
from io import BytesIO
import os
import smtplib
from email.message import EmailMessage

import streamlit as st
from openpyxl import Workbook, load_workbook
from zipfile import BadZipFile

APP_DIR = Path(__file__).parent
DATA_FILE = APP_DIR / "daily_expenses.xlsx"
RECEIPT_DIR = APP_DIR / "receipts"
RECEIPT_DIR.mkdir(exist_ok=True)

HEADERS = ["Timestamp", "Date", "Time", "Category", "Expense Type", "Vendor/Notes", "Amount", "Receipt File"]
EXPENSE_TYPES = ["Fuel", "Meals", "Office Supplies", "Travel", "Software", "Utilities", "Other"]

st.set_page_config(page_title="Daily Expense Tracker", page_icon="💳", layout="centered")

st.markdown("""
<style>
.stApp {background: linear-gradient(180deg,#eef2ff,#ffffff);} 
.block-container {max-width:440px; padding-top:1rem; padding-bottom:5rem;}
.hero {background:linear-gradient(135deg,#111827,#2563eb); color:white; padding:24px; border-radius:28px; margin-bottom:16px; box-shadow:0 18px 45px rgba(15,23,42,.15);} 
.card {background:white; padding:20px; border-radius:24px; border:1px solid #e5e7eb; box-shadow:0 10px 28px rgba(15,23,42,.08); margin-bottom:14px;}
.metric-card {background:#f8fafc; border:1px solid #e2e8f0; border-radius:18px; padding:14px; margin-bottom:10px;}
.metric-card small {color:#64748b;}
.metric-card b {font-size:24px; color:#111827;}
</style>
""", unsafe_allow_html=True)


def create_new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense_Log"
    ws.append(HEADERS)
    ws2 = wb.create_sheet("Dashboard")
    ws2["A1"] = "Daily Expense Tracker Dashboard"
    ws2["A3"] = "Use the app dashboard for live totals."
    wb.save(DATA_FILE)


def ensure_workbook():
    # Create the Excel file if it is missing. If GitHub/Streamlit has an
    # empty, corrupted, or non-Excel daily_expenses.xlsx file, delete it and
    # rebuild a clean workbook instead of crashing with zipfile.BadZipFile.
    if not DATA_FILE.exists() or DATA_FILE.stat().st_size < 100:
        create_new_workbook()
        return
    try:
        wb = load_workbook(DATA_FILE)
        if "Expense_Log" not in wb.sheetnames:
            DATA_FILE.unlink(missing_ok=True)
            create_new_workbook()
    except (BadZipFile, OSError, KeyError, ValueError):
        DATA_FILE.unlink(missing_ok=True)
        create_new_workbook()


def get_rows():
    ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb["Expense_Log"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            rows.append(dict(zip(HEADERS, row)))
    return rows


def append_expense(category, expense_type, notes, amount, receipt_name):
    ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb["Expense_Log"]
    now = datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d %H:%M:%S"),
        now.strftime("%Y-%m-%d"),
        now.strftime("%I:%M %p"),
        category,
        expense_type,
        notes,
        float(amount),
        receipt_name,
    ])
    wb.save(DATA_FILE)


def save_receipt(file, category):
    if file is None:
        return ""
    safe = file.name.replace(" ", "_")
    name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{category}_{safe}"
    path = RECEIPT_DIR / name
    path.write_bytes(file.getbuffer())
    return name


def workbook_bytes():
    ensure_workbook()
    return DATA_FILE.read_bytes()


def send_email_now():
    smtp_host = st.secrets.get("SMTP_HOST", "")
    smtp_port = int(st.secrets.get("SMTP_PORT", 587))
    smtp_user = st.secrets.get("SMTP_USER", "")
    smtp_password = st.secrets.get("SMTP_PASSWORD", "")
    email_to = st.secrets.get("EMAIL_TO", "kevin@unlimited-cyber.com")
    if not smtp_host or not smtp_user or not smtp_password:
        raise ValueError("Email secrets are missing. Add SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, and EMAIL_TO in Streamlit Secrets.")
    msg = EmailMessage()
    msg["Subject"] = "Daily Expense Spreadsheet"
    msg["From"] = smtp_user
    msg["To"] = email_to
    msg.set_content("Attached is the latest daily expense spreadsheet.")
    msg.add_attachment(workbook_bytes(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="daily_expenses.xlsx")
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


st.markdown('<div class="hero"><h1>Daily Expense Tracker</h1><p>Track Personal and Business spending with receipt backup.</p></div>', unsafe_allow_html=True)

pin = st.secrets.get("APP_PIN", "1234")
if "unlocked" not in st.session_state:
    st.session_state.unlocked = False

if not st.session_state.unlocked:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    entered = st.text_input("Enter PIN", type="password")
    if st.button("Unlock", type="primary", use_container_width=True):
        if entered == pin:
            st.session_state.unlocked = True
            st.rerun()
        else:
            st.error("Incorrect PIN")
    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

tab1, tab2, tab3 = st.tabs(["Add", "Dashboard", "Export"])

with tab1:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    category = st.radio("Category", ["Personal", "Business"], horizontal=True)
    amount = st.number_input("Amount", min_value=0.0, step=0.01, format="%.2f")
    expense_type = st.selectbox("Expense type", EXPENSE_TYPES)
    notes = st.text_input("Vendor / Notes")
    receipt = st.file_uploader("Receipt photo optional", type=["png", "jpg", "jpeg", "pdf"])
    if st.button("Submit Expense", type="primary", use_container_width=True):
        if amount <= 0:
            st.error("Enter an amount greater than $0.")
        else:
            receipt_name = save_receipt(receipt, category)
            append_expense(category, expense_type, notes, amount, receipt_name)
            st.success("Expense saved to spreadsheet.")
    st.markdown('</div>', unsafe_allow_html=True)

with tab2:
    rows = get_rows()
    personal = sum(float(r.get("Amount") or 0) for r in rows if r.get("Category") == "Personal")
    business = sum(float(r.get("Amount") or 0) for r in rows if r.get("Category") == "Business")
    st.markdown(f'<div class="metric-card"><small>Personal Total</small><br><b>${personal:,.2f}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="metric-card"><small>Business Total</small><br><b>${business:,.2f}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="metric-card"><small>All Expenses</small><br><b>${personal+business:,.2f}</b></div>', unsafe_allow_html=True)
    st.write("Recent entries")
    st.dataframe(list(reversed(rows[-20:])), use_container_width=True)

with tab3:
    st.download_button("Download Excel Spreadsheet", data=workbook_bytes(), file_name="daily_expenses.xlsx", use_container_width=True)
    if st.button("Send Test Email", use_container_width=True):
        try:
            send_email_now()
            st.success("Email sent.")
        except Exception as e:
            st.error(str(e))
    st.caption("For automatic 11:59 PM email, use GitHub Actions or a paid hosted scheduler. Streamlit Cloud apps can sleep, so in-app scheduling is not reliable.")
